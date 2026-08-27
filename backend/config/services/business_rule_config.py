import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.conf import settings
from openpyxl import load_workbook


@dataclass(frozen=True)
class NominatorRuleConfig:
    gross_amount: dict[str, int | None]
    tax_type: dict[str, str | None]
    payment_clauses: dict[str, dict[str, dict[str, str]]]
    domestic_country_codes: set[str]
    tax_rates: dict[str, dict[str, float]]


@dataclass(frozen=True)
class BusinessRuleConfig:
    nominator: NominatorRuleConfig
    source: str


def load_business_rule_config() -> BusinessRuleConfig:
    try:
        return load_business_rule_config_from_workbook(
            settings.BUSINESS_RULE_WORKBOOK,
            settings.BUSINESS_RULE_SHEET,
        )
    except (FileNotFoundError, ValueError):
        return load_business_rule_config_from_json(settings.BUSINESS_RULE_CONFIG_PATH)


def load_business_rule_config_from_json(config_path: Path) -> BusinessRuleConfig:
    path = config_path or settings.BUSINESS_RULE_CONFIG_PATH
    with path.open(encoding="utf-8") as config_file:
        raw_config = json.load(config_file)
    raw_config["_source"] = f"json:{path}"
    return parse_business_rule_config(raw_config)


def load_business_rule_config_from_workbook(workbook_path: Path, sheet_name: str) -> BusinessRuleConfig:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    raw_config = parse_rules_sheet_to_raw_config(workbook[sheet_name])
    raw_config["_source"] = f"xlsx:{workbook_path}#{sheet_name}"
    return parse_business_rule_config(raw_config)


def parse_rules_sheet_to_raw_config(sheet) -> dict[str, Any]:
    raw_config: dict[str, Any] = {
        "tax_type": {},
        "domestic_country_codes": [],
        "contract_type": {},
    }
    current_section = ""
    current_tax_region = ""
    current_contract_type = ""
    current_contract_region = ""
    current_contract_key = ""

    for row in sheet.iter_rows(values_only=True):
        cells = [_clean_cell(value) for value in row]
        if not any(cells):
            continue

        if cells[0]:
            current_section = _normalize_key(cells[0])
            current_tax_region = ""
            current_contract_type = ""
            current_contract_region = ""
            continue

        if current_section == "tax_type":
            if cells[1]:
                current_tax_region = _normalize_region(cells[1])
                raw_config["tax_type"].setdefault(current_tax_region, {})
                continue
            if current_tax_region and cells[2]:
                raw_config["tax_type"][current_tax_region][_normalize_tax_type(cells[2])] = cells[3]
            continue

        if current_section == "domestic_country_codes":
            if cells[1]:
                raw_config["domestic_country_codes"].append(str(cells[1]).strip())
            continue

        if current_section == "contract_type":
            if cells[1]:
                current_contract_type = _normalize_contract_type(cells[1])
                raw_config["contract_type"].setdefault(current_contract_type, {})
                current_contract_region = ""
                current_contract_key = ""
                continue
            if current_contract_type and cells[2]:
                current_contract_region = _normalize_region(cells[2])
                raw_config["contract_type"][current_contract_type].setdefault(
                    current_contract_region,
                    {},
                )
                current_contract_key = ""
                continue
            if current_contract_type and current_contract_region and cells[3]:
                key = _normalize_key(cells[3])
                current_contract_key = key
                if cells[4] not in (None, ""):
                    raw_config["contract_type"][current_contract_type][current_contract_region][key] = cells[4]
                else:
                    raw_config["contract_type"][current_contract_type][current_contract_region].setdefault(key, {})
                continue
            if current_contract_type and current_contract_region and current_contract_key and cells[4]:
                language = _normalize_language(cells[4])
                raw_config["contract_type"][current_contract_type][current_contract_region].setdefault(
                    current_contract_key,
                    {},
                )[language] = cells[5]

    return raw_config


def parse_business_rule_config(raw_config: dict[str, Any]) -> BusinessRuleConfig:
    contract_types = raw_config.get("contract_type") or raw_config.get("contract-type", {})
    nominator = contract_types.get("nominator", {})
    domestic_nominator = nominator.get("domestic", {})
    overseas_nominator = nominator.get("overseas", {})
    return BusinessRuleConfig(
        nominator=NominatorRuleConfig(
            gross_amount={
                "domestic": _optional_int(domestic_nominator.get("gross_amount")),
                "overseas": _optional_int(overseas_nominator.get("gross_amount")),
            },
            tax_type={
                "domestic": _normalize_tax_type(domestic_nominator.get("tax_type")),
                "overseas": _normalize_tax_type(overseas_nominator.get("tax_type")),
            },
            payment_clauses={
                "domestic": _extract_payment_clauses(domestic_nominator),
                "overseas": _extract_payment_clauses(overseas_nominator),
            },
            domestic_country_codes={
                str(value).strip()
                for value in raw_config.get("domestic_country_codes", [])
                if str(value).strip()
            },
            tax_rates={
                _normalize_region(region): {
                    _normalize_tax_type(tax_type): float(rate or 0)
                    for tax_type, rate in rules.items()
                }
                for region, rules in raw_config.get("tax_type", {}).items()
            },
        ),
        source=str(raw_config.get("_source", "unknown")),
    )


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def _extract_payment_clauses(values: dict[str, Any]) -> dict[str, str]:
    clauses: dict[str, dict[str, str]] = {"kor": {}, "eng": {}}
    for key, value in values.items():
        if not key.startswith("payment_clause") or value in (None, ""):
            continue
        if isinstance(value, dict):
            for language, text in value.items():
                normalized_language = _normalize_language(language)
                if text not in (None, ""):
                    clauses.setdefault(normalized_language, {})[key] = str(text)
        else:
            clauses["kor"][key] = str(value)
            clauses["eng"][key] = str(value)
    return clauses


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        if not value or value.startswith("#"):
            return None
        return value
    return value


def _normalize_key(value: Any) -> str:
    return str(value or "").strip().replace("-", "_")


def _normalize_region(value: Any) -> str:
    normalized = _normalize_key(value)
    if normalized in {"domestic", "overseas"}:
        return normalized
    return normalized


def _normalize_contract_type(value: Any) -> str:
    return _normalize_key(value).strip()


def _normalize_tax_type(value: Any) -> str | None:
    if value is None:
        return None
    normalized = _normalize_key(value)
    aliases = {
        "other_income": "other_income",
        "tax_exampt": "tax_exempt",
        "tax_exempt": "tax_exempt",
    }
    return aliases.get(normalized, normalized)


def _normalize_language(value: Any) -> str:
    normalized = _normalize_key(value)
    if normalized in {"eng", "en", "english"}:
        return "eng"
    return "kor"
