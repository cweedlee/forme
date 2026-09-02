from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from config.services.project_settings import ProjectRuntimeConfig, load_project_config

SEOUL_TZ = ZoneInfo("Asia/Seoul")


@dataclass(frozen=True)
class TableRow:
    data_key: str
    source_row: int
    values: dict[str, Any]


@dataclass(frozen=True)
class TableData:
    sheet_name: str
    keys: dict[str, str]
    data: list[TableRow]
    metadata: dict[str, Any]


def load_table(sheet_key: str, workbook_path: Path | None = None) -> TableData:
    config = load_project_config()
    return Workbook(workbook_path or config.workbook_path, config=config).get_sheet(sheet_key)


class Workbook:
    """Read project sheets using the single, project-wide table layout."""

    def __init__(self, workbook_path: Path, *, config: ProjectRuntimeConfig | None = None) -> None:
        self.workbook_path = Path(workbook_path)
        self.config = config or load_project_config()

    def get_sheet(self, sheet_key: str) -> TableData:
        try:
            sheet_name = self.config.sheets[sheet_key]
        except KeyError as exc:
            raise ValueError(f"프로젝트 설정에 sheet가 없습니다: {sheet_key}") from exc
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {sheet_name}")
            return self._parse_sheet(sheet_key, sheet_name, workbook[sheet_name])
        finally:
            workbook.close()

    def _parse_sheet(self, sheet_key: str, sheet_name: str, sheet: Any) -> TableData:
        layout = self.config.data_sheet_layout
        data_row = int(layout["data_row"])
        data_column = int(layout["data_column"])
        name_row = int(layout["name_row"])
        var_row = int(layout["var_row"])
        data_key_name = str(layout["data_key"])

        keys: dict[str, str] = {}
        column_indexes: dict[str, int] = {}
        for column in range(data_column, sheet.max_column + 1):
            var_name = _clean_text(sheet.cell(row=var_row, column=column).value)
            if not var_name:
                continue
            if var_name in column_indexes:
                raise ValueError(f"{sheet_name} 시트에 중복된 var 헤더가 있습니다: {var_name}")
            display_name = _clean_text(sheet.cell(row=name_row, column=column).value) or var_name
            keys[var_name] = display_name
            column_indexes[var_name] = column

        if not keys:
            raise ValueError(f"{sheet_name} 시트의 var 헤더 행이 비어 있습니다: {var_row}")
        if data_key_name not in column_indexes:
            raise ValueError(f"{sheet_name} 시트에 data-key var 헤더가 없습니다: {data_key_name}")

        rows: list[TableRow] = []
        seen_data_keys: set[str] = set()
        for row_number in range(data_row, sheet.max_row + 1):
            values = {
                var_name: _clean_cell(sheet.cell(row=row_number, column=column).value)
                for var_name, column in column_indexes.items()
            }
            if not _has_any_value(values.values()):
                continue
            data_key = _clean_text(values[data_key_name])
            if not data_key:
                continue
            if data_key in seen_data_keys:
                raise ValueError(f"{sheet_name} 시트의 data-key가 중복되었습니다: {data_key}")
            seen_data_keys.add(data_key)
            rows.append(TableRow(data_key=data_key, source_row=row_number, values=values))

        stat = self.workbook_path.stat()
        return TableData(
            sheet_name=sheet_name,
            keys=keys,
            data=rows,
            metadata={
                "data_row": data_row,
                "name_row": name_row,
                "var_row": var_row,
                "data_key": data_key_name,
                "template_path_kor": str(self.config.template_for(sheet_key, "kor").path),
                "template_path_eng": str(self.config.template_for(sheet_key, "eng").path),
                "workbook_mtime": datetime.fromtimestamp(stat.st_mtime, tz=SEOUL_TZ).isoformat(),
                "source_mode": "xlsx:project-sheet",
            },
        )


def _clean_cell(value: Any) -> Any:
    return value.strip() if isinstance(value, str) else value


def _clean_text(value: Any) -> str:
    return str(_clean_cell(value) or "").strip()


def _has_any_value(values: Any) -> bool:
    return any(value not in (None, "") for value in values)
