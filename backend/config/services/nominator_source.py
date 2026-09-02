from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from config.services.person_rows import (
    PersonSourceRow,
    field_key_from_header,
    is_person_code_header_row,
    parse_person_source_row,
)
from config.services.project_settings import load_project_config


FIRST_COLUMN = 1
CODE_HEADER_ROW = 2
DATA_FIRST_ROW = 3
ENGAGEMENT_TYPE = "nominator"
DECISION_COLUMNS = ["계약서 상태"]
SEOUL_TZ = ZoneInfo("Asia/Seoul")
REQUIRED_CONTEXT_FIELDS = {
    "key": {"식별자", "key"},
    "person_name_kor": {"국문 성명", "국문 이름", "성명"},
    "tax_residence": {"세법상 거주국", "국가"},
    "work_location": {"업무수행장소"},
    "gross_amount": {"계약금액"},
    "income_type": {"소득종류"},
    "tax_rate": {"원천징수율"},
    "tax_amount": {"원천징수세액(KRW)"},
    "final_amount": {"최종 지급액"},
}


@dataclass(frozen=True)
class NominatorWorkbookTable:
    workbook_path: Path
    sheet_name: str
    columns: list[str]
    field_keys: list[str | None]
    rows: list[dict[str, Any]]
    decision_columns: list[str]
    metadata: dict[str, Any]


def load_nominator_table(
    workbook_path: Path,
    sheet_name: str | None = None,
) -> NominatorWorkbookTable:
    project_config = load_project_config()
    selected_sheet_name = sheet_name or project_config.people_sheet_for("nominator")
    layout = project_config.people_layout_for("nominator")
    sheet = load_people_sheet(workbook_path, selected_sheet_name)
    headers = read_code_headers(sheet, row_number=layout["code_header_row"])
    nature_headers = read_row_values(sheet, layout["nature_header_row"])
    visible_indexes = visible_column_indexes(headers, layout.get("visible_columns", []))
    columns = display_columns(headers, nature_headers, visible_indexes)
    field_keys = visible_field_keys(headers, visible_indexes)
    source_rows = read_non_empty_rows(sheet, first_row=layout["data_first_row"])
    rows = build_nominator_table_rows(
        source_rows,
        headers,
        visible_indexes,
        selected_sheet_name,
        layout["data_first_row"],
    )

    return NominatorWorkbookTable(
        workbook_path=workbook_path,
        sheet_name=selected_sheet_name,
        columns=columns,
        field_keys=field_keys,
        rows=rows,
        decision_columns=DECISION_COLUMNS,
        metadata=build_table_metadata(workbook_path),
    )


def load_nominator_person_from_workbook(
    workbook_path: Path,
    source_row: int,
    sheet_name: str | None = None,
) -> PersonSourceRow | None:
    project_config = load_project_config()
    selected_sheet_name = sheet_name or project_config.people_sheet_for("nominator")
    layout = project_config.people_layout_for("nominator")
    sheet = load_people_sheet(workbook_path, selected_sheet_name)
    headers = read_code_headers(sheet, row_number=layout["code_header_row"])

    for row_number, values in read_non_empty_rows(sheet, first_row=layout["data_first_row"]):
        if row_number == source_row:
            return parse_person_source_row(
                source_row=row_number,
                headers=headers,
                values=values,
                engagement_type=ENGAGEMENT_TYPE,
            )
    return None


def load_people_sheet(workbook_path: Path, sheet_name: str):
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    return workbook[sheet_name]


def read_code_headers(sheet, *, row_number: int = CODE_HEADER_ROW) -> list[str]:
    values = read_row_values(sheet, row_number)
    if not is_person_code_header_row(values):
        labels = ", ".join(str(value or "").strip() for value in values if value)
        raise ValueError(f"{row_number}행 코드 헤더가 올바르지 않습니다: {labels}")
    return [str(value or "").strip() for value in values]


def read_non_empty_rows(sheet, *, first_row: int) -> list[tuple[int, list[Any]]]:
    rows: list[dict[str, Any]] = []
    for row_number in range(first_row, sheet.max_row + 1):
        values = read_row_values(sheet, row_number)
        if _has_any_value(values):
            rows.append((row_number, values))
    return rows


def read_row_values(sheet, row_number: int) -> list[Any]:
    return [
        _clean_cell(sheet.cell(row=row_number, column=column).value)
        for column in range(FIRST_COLUMN, sheet.max_column + 1)
    ]


def build_table_metadata(workbook_path: Path) -> dict[str, Any]:
    stat = workbook_path.stat()
    return {
        "workbook_mtime": datetime.fromtimestamp(stat.st_mtime, tz=SEOUL_TZ).isoformat(),
        "source_mode": "xlsx:nominator-sheet",
    }


def build_nominator_table_rows(
    source_rows: list[tuple[int, list[Any]]],
    headers: list[str],
    visible_indexes: list[int],
    sheet_name: str,
    data_first_row: int = DATA_FIRST_ROW,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for row_number, values in source_rows:
        visible_values = filter_values(values, visible_indexes)
        if not _has_any_value(visible_values):
            continue

        if row_number < data_first_row:
            rows.append(_build_preview_row(row_number, visible_values))
            continue

        person = None
        person = parse_person_source_row(
            source_row=row_number,
            headers=headers,
            values=values,
            engagement_type=ENGAGEMENT_TYPE,
        )
        if not _has_person_identity(person):
            continue
        rows.append(_build_preview_row(row_number, visible_values, person))

    return rows


def visible_column_indexes(headers: list[str], visible_columns: list[str]) -> list[int]:
    visible_keys = {
        field_key_from_header(column)
        for column in visible_columns
        if field_key_from_header(column) is not None
    }
    indexes = [
        index
        for index, header in enumerate(headers)
        if header and (not visible_keys or field_key_from_header(header) in visible_keys)
    ]
    return indexes


def display_columns(headers: list[str], nature_headers: list[Any], visible_indexes: list[int]) -> list[str]:
    columns: list[str] = []
    for index in visible_indexes:
        nature_label = nature_headers[index] if index < len(nature_headers) else None
        code_label = headers[index] if index < len(headers) else ""
        columns.append(str(nature_label or code_label or "").strip())
    return columns


def visible_field_keys(headers: list[str], visible_indexes: list[int]) -> list[str | None]:
    return [
        field_key_from_header(headers[index]) if index < len(headers) else None
        for index in visible_indexes
    ]


def filter_values(values: list[Any], visible_indexes: list[int]) -> list[Any]:
    return [
        values[index] if index < len(values) else None
        for index in visible_indexes
    ]


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _has_any_value(values: list[Any]) -> bool:
    return any(value not in (None, "") for value in values)


def _has_person_identity(person: PersonSourceRow) -> bool:
    return bool(person.key or person.name.kor or person.residence_country)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().replace("-", "_").lower()


def _build_preview_row(
    row_number: int,
    values: list[Any],
    person: PersonSourceRow | None = None,
) -> dict[str, Any]:
    row = {
        "source_row": row_number,
        "values": values,
        "decisions": {},
        "person": _serialize_person(person) if person else None,
    }
    if person:
        row["decisions"] = build_row_decisions(person)
    return row


def build_row_decisions(person: PersonSourceRow) -> dict[str, str]:
    missing = missing_required_context_fields(person)
    if missing:
        return {
            "계약서 상태": "필수값 누락: " + ", ".join(missing),
        }
    return {
        "계약서 상태": "준비",
    }


def missing_required_context_fields(person: PersonSourceRow) -> list[str]:
    missing = []
    for field_name, aliases in REQUIRED_CONTEXT_FIELDS.items():
        if _mapped_value(person.raw_by_header, aliases) in (None, ""):
            missing.append(field_name)
    return missing


def _mapped_value(row: dict[str, Any], aliases: set[str]) -> Any:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for header, value in row.items():
        if _normalize_header(header) in normalized_aliases:
            return value
    return None


def _format_krw(value: Any) -> str:
    if value is None:
        return ""
    return f"{value:,}원"


def _serialize_person(person: PersonSourceRow | None) -> dict[str, Any] | None:
    if not person:
        return None
    return {
        "source_row": person.source_row,
        "engagement_type": person.engagement_type,
        "name": {
            "kor": person.name.kor,
            "eng": person.name.eng,
        },
        "country_code": person.country_code,
        "residence_country": person.residence_country,
        "workplace": person.workplace,
        "key": person.key,
        "raw_by_header": person.raw_by_header,
    }
