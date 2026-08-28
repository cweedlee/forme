from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config.services.project_settings import load_project_config


COUNTRY_NAME_HEADERS = {
    "country",
    "country_name",
    "name",
    "residence_country",
    "tax_residence",
    "국가",
    "국가명",
    "국문명",
    "한글명",
    "거주국",
    "세법상 거주국",
}
COUNTRY_CODE_HEADERS = {"country_code", "code", "iso", "iso_code", "국가코드", "코드"}


def resolve_country_code(country_name: str, workbook_path: Path | None = None) -> str:
    country_codes = load_country_code_map(workbook_path)
    normalized_name = _normalize(country_name)
    try:
        return country_codes[normalized_name]
    except KeyError as exc:
        raise ValueError(f"country-code 시트에서 국가코드를 찾을 수 없습니다: {country_name}") from exc


def load_country_code_map(workbook_path: Path | None = None) -> dict[str, str]:
    project_config = load_project_config()
    path = workbook_path or project_config.workbook_path
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = project_config.country_code_sheet
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"country-code 시트를 찾을 수 없습니다: {sheet_name}")

    rows = [
        [cell for cell in row]
        for row in workbook[sheet_name].iter_rows(values_only=True)
        if any(cell not in (None, "") for cell in row)
    ]
    return parse_country_code_rows(rows)


def parse_country_code_rows(rows: list[list[Any]]) -> dict[str, str]:
    if not rows:
        raise ValueError("country-code 시트가 비어 있습니다.")

    header_row_index, name_index, code_index = find_country_code_header(rows)
    country_codes: dict[str, str] = {}
    for row in rows[header_row_index + 1 :]:
        name = _cell(row, name_index)
        code = _cell(row, code_index)
        if name and code:
            country_codes[_normalize(name)] = str(code).strip()

    if not country_codes:
        raise ValueError("country-code 시트에서 국가코드 값을 찾을 수 없습니다.")
    return country_codes


def find_country_code_header(rows: list[list[Any]]) -> tuple[int, int, int]:
    for row_index, row in enumerate(rows[:10]):
        normalized = [_normalize_header(value) for value in row]
        name_index = _first_index(normalized, COUNTRY_NAME_HEADERS)
        code_index = _first_index(normalized, COUNTRY_CODE_HEADERS)
        if name_index is not None and code_index is not None:
            return row_index, name_index, code_index
    raise ValueError("country-code 시트에 국가명/국가코드 헤더가 없습니다.")


def _first_index(values: list[str], accepted: set[str]) -> int | None:
    normalized_accepted = {_normalize_header(value) for value in accepted}
    for index, value in enumerate(values):
        if value in normalized_accepted:
            return index
    return None


def _cell(row: list[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def _normalize(value: Any) -> str:
    return str(value or "").strip()


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().replace("-", "_").lower()
